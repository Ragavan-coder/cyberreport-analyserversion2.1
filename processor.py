import re
import pdfplumber
from dateutil import parser
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# =====================================================
# FIXED FIELD SCHEMA (ONLY ORIGINAL DATA)
# =====================================================
FIELDS = [
    "Complaint ID",
    "Date Filed",
    "Date Accepted",
    "Time Accepted",
    "Complainant Name",
    "Email",
    "Mobile Number",
    "State",
    "District",
    "Cybercrime Type",
    "Sub-Category",
    "Platform",
    "Amount Lost",
    "Complaint Status",
    "FIR Status",
    "Investigation Status"
]

# =====================================================
# LABEL MAP (STRICT – NO ASSUMPTIONS)
# =====================================================
LABELS = {
    "Complaint ID": [r"acknowledg(e)?ment\s*number", r"complaint\s*id"],
    "Date Filed": [r"complaint\s*date"],
    "Cybercrime Type": [r"category\s*of\s*complaint"],
    "Sub-Category": [r"sub\s*category\s*of\s*complaint"],
    "Complainant Name": [r"name\s*:"],
    "Email": [r"email"],
    "Mobile Number": [r"mobile"],
    "District": [r"district"],
    "State": [r"state"],
    "Amount Lost": [r"total\s*fraudulent\s*amount"]
}

# =====================================================
# PDF TEXT EXTRACTION
# =====================================================
def extract_text_from_pdf(path):
    text = ""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text += "\n" + t
    return text

# =====================================================
# SAFE COMPLAINT SPLITTER (BUG FIXED)
# =====================================================
def split_complaints(text):
    if not text:
        return []

    blocks = re.split(
        r"(?=Complaint\s*(?:ID|Id|No)|Complaint\s*Type\s*:)",
        text,
        flags=re.I
    )

    cleaned = []
    for b in blocks:
        if isinstance(b, str):
            b = b.strip()
            if len(b) > 200:
                cleaned.append(b)

    return cleaned

# =====================================================
# NORMALIZATION
# =====================================================
def clean_value(val):
    if not val:
        return ""
    return re.sub(r"\s+", " ", val).strip(" :-")

def normalize_date(val):
    try:
        return parser.parse(val, dayfirst=True).strftime("%d/%m/%Y")
    except:
        return ""

def normalize_amount(val):
    if not val:
        return ""
    num = re.sub(r"[^\d]", "", val)
    return f"₹{int(num):,}" if num.isdigit() else ""

# =====================================================
# FIELD EXTRACTION (PER COMPLAINT)
# =====================================================
def extract_fields(block):
    data = {f: "NULL" for f in FIELDS}
    lines = [l.strip() for l in block.splitlines() if l.strip()]

    for line in lines:
        low = line.lower()

        for field, patterns in LABELS.items():
            for p in patterns:
                if re.search(p, low):
                    value = line.split(":", 1)[-1]
                    data[field] = clean_value(value)

    # Email fallback
    if data["Email"] == "NULL":
        m = re.search(r"[\w\.-]+@[\w\.-]+\.\w+", block)
        if m:
            data["Email"] = m.group(0)

    # Accepted date & time
    m = re.search(
        r"complaint\s*accepted\s*date\s*[:\-]?\s*(\d{1,2}/\d{1,2}/\d{4})\s*(\d{1,2}:\d{2}:\d{2}\s*[AP]M)",
        block,
        re.I
    )
    if m:
        data["Date Accepted"] = normalize_date(m.group(1))
        data["Time Accepted"] = m.group(2)

    data["Date Filed"] = normalize_date(data["Date Filed"])
    data["Amount Lost"] = normalize_amount(data["Amount Lost"])

    # Platform
    if "UPI" in block.upper():
        data["Platform"] = "UPI"
    else:
        data["Platform"] = "Other"

    # Status
    T = block.upper()
    data["Complaint Status"] = (
        "CLOSED" if "CLOSED" in T else
        "UNDER PROCESS" if "UNDER PROCESS" in T else
        "ACCEPTED" if "COMPLAINT ACCEPTED" in T else
        "PENDING"
    )

    data["FIR Status"] = "FILED" if "FIR" in T else "NOT FILED"

    data["Investigation Status"] = (
        "CLOSED" if "CLOSED" in T else
        "ONGOING" if "UNDER PROCESS" in T else
        "NOT STARTED"
    )

    return data

# =====================================================
# REMOVE DUPLICATES (CRITICAL)
# =====================================================
def deduplicate(records):
    seen = set()
    unique = []

    for r in records:
        key = (
            r.get("Complaint ID"),
            r.get("Date Filed"),
            r.get("Complainant Name"),
            r.get("Amount Lost")
        )
        if key not in seen:
            seen.add(key)
            unique.append(r)

    return unique

# =====================================================
# MAIN PROCESS PDF
# =====================================================
def process_pdf(pdf_path):
    text = extract_text_from_pdf(pdf_path)
    blocks = split_complaints(text)

    records = []
    for block in blocks:
        record = extract_fields(block)
        records.append(record)

    records = deduplicate(records)
    return records

# =====================================================
# EXCEL FORMATTING
# =====================================================
def format_worksheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    ws.freeze_panes = "A2"

# =====================================================
# SAVE EXCEL
# =====================================================
def save_consolidated_excel(all_records, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cyber Complaints"

    ws.append(FIELDS)

    for record in all_records:
        ws.append([record.get(f, "NULL") for f in FIELDS])

    format_worksheet(ws)
    wb.save(out_path)
